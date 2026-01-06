#!/usr/bin/env python3
"""
Script to create Excel templates for customer import
- Template 1: Cash customers (Khách bộ - INTERNAL)
- Template 2: Credit customers (Khách công nợ - EXTERNAL)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template(filename, title, customer_type, include_credit_limit=False):
    """Create an Excel template for customer import"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Khách hàng"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="4472C4")
    instruction_font = Font(italic=True, size=10, color="C00000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = title
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Instructions
    instruction_text = f"Hướng dẫn: Điền thông tin khách hàng từ dòng 5 trở đi. Loại khách hàng: {customer_type}. Cột có dấu (*) bắt buộc."
    ws['A2'] = instruction_text
    ws['A2'].font = instruction_font
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(wrap_text=True)

    # Hidden customer type value for processing
    ws['A3'] = customer_type
    ws['A3'].font = Font(size=8, color="999999")

    # Headers
    if include_credit_limit:
        headers = [
            "Mã KH",
            "Tên khách hàng (*)",
            "Mã số thuế",
            "Địa chỉ",
            "Số điện thoại (*)",
            "Hạn mức công nợ",
            "Ghi chú"
        ]
        sample_data = [
            ["", "Công ty TNHH ABC", "0123456789", "123 Đường ABC, TP.HCM", "0901234567", "50000000", "Khách hàng VIP"],
            ["", "Công ty CP XYZ", "9876543210", "456 Đường XYZ, Hà Nội", "0912345678", "30000000", "Khách hàng thường xuyên"],
            ["", "Cửa hàng DEF", "", "789 Đường DEF, Đà Nẵng", "0923456789", "20000000", ""],
        ]
        column_widths = [15, 35, 15, 40, 18, 20, 30]
    else:
        headers = [
            "Mã KH",
            "Tên khách hàng (*)",
            "Mã số thuế",
            "Địa chỉ",
            "Số điện thoại (*)",
            "Ghi chú"
        ]
        sample_data = [
            ["", "Khách lẻ Nguyễn Văn A", "", "123 Đường ABC, TP.HCM", "0901234567", "Khách hàng nội bộ"],
            ["", "Khách bộ Trần Thị B", "", "456 Đường XYZ, Hà Nội", "0912345678", ""],
            ["", "Cá nhân C", "", "789 Đường DEF, Đà Nẵng", "0923456789", "Khách quen"],
        ]
        column_widths = [15, 35, 15, 40, 18, 30]

    # Write headers (row 4)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Write sample data
    for row_idx, row_data in enumerate(sample_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    # Set column widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 35

    # Save file
    wb.save(filename)
    print(f"✓ Đã tạo: {filename}")

# Create templates
output_dir = '/home/seth/WorkSpace/SWP/mau so/'

# Template 1: Credit customers (Khách công nợ)
create_template(
    filename=f'{output_dir}MauImportKhachCongNo.xlsx',
    title='MẪU IMPORT KHÁCH HÀNG CÔNG NỢ',
    customer_type='EXTERNAL',
    include_credit_limit=True
)

# Template 2: Cash customers (Khách bộ)
create_template(
    filename=f'{output_dir}MauImportKhachBo.xlsx',
    title='MẪU IMPORT KHÁCH BỘ (KHÁCH LẺ)',
    customer_type='INTERNAL',
    include_credit_limit=False
)

print("\n✓ Hoàn thành! Đã tạo 2 template Excel.")
